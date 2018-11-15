import re
import os
import pymongo
import openpyxl
from pathlib import Path
from bson import ObjectId
import copy

'''
Prog will iterate through all sheet in excel file that is placed in 'placeWholesaleExcelFileHere'.
It will parse through the table and insert the desired cols into mongodb.

Desired header names names will be transformed to lowercased alphanumeric value.
    ex) Product No. -> productno
Desired transformed header names must be equal across all csv's.
    ie) The following is fine because they are transformed to the same thing
        File1: Product No.  -> productno
        File2: product no   -> productno
The column order of the header names for a excel file do not matter.

Coder Note:
    At the moment, ship method excel must have columns in certain order. Ie) UPC first col...
'''

def getLoweredAlphaNumericStr(string:str)->str:
    if type(string) is not str:
        return None
    string = re.sub('[^A-Za-z0-9]+', '', string)
    string = str.lower(string)
    return string

#Will remove any non-alphaNumeric character in search
def getColForHeaderName(sheet, headerRow:int, headerName:str)->int:
    max_column = sheet.max_column
    headerNameOfOnlyLetterAndNum = getLoweredAlphaNumericStr(headerName)

    for i in range(1, max_column+1):
        cellVal = sheet.cell(row=headerRow, column=i).value
        if(cellVal == None):
            continue
        headerCellValueOfOnlyLetterAndNum = getLoweredAlphaNumericStr(cellVal)
        if headerCellValueOfOnlyLetterAndNum == None:
            return -1
        if(headerNameOfOnlyLetterAndNum in headerCellValueOfOnlyLetterAndNum):
            return i
    
def getColNumToHeaderNameDict(sheet, headerRow:int, headersToMongoNameDict:dict)->dict:
    colNumToHeaderNameDict = {}
    for headerName in headersToMongoNameDict:
        headerCol = getColForHeaderName(sheet, headerRow, headerName)
        if headerCol == -1:
            continue
        colNumToHeaderNameDict[headerCol] = headerName
    
    return colNumToHeaderNameDict

'''
Returns true if in the form of '$#.##'
ex) isMoneyCellVal('$4.2') == true  isMoneyCellVal('4.2') == false
ex) isMoneyCellVal('$.2') == true  isMoneyCellVal('abc') == false
'''
def isMoneyCellVal(val:str)->bool:
    return val[0] == '$' and eval(val[1:])

def getDictMongoHeaderToCellVal(row:list, headersToMongoNameDict:dict
                          , colNumToHeaderNameDict:dict) ->dict:
    retDict = {}

    i = 1
    for cell in row:
        cellVal = cell.value
        if i in colNumToHeaderNameDict and cellVal != '':
            if isinstance(cellVal, str) and isMoneyCellVal(cellVal):
                cellVal = cellVal[1:] #Trims the '$' in front

            csvHeaderName = colNumToHeaderNameDict[i]
            mongoDbName = headersToMongoNameDict[csvHeaderName]
            retDict[mongoDbName] = cellVal
        i = i+1
    return retDict

def getShippingInfo(upcToShippingInfoDict, upc):
    if upc in upcToShippingInfoDict:
        terminateIndex = 0
        shippingInfos = upcToShippingInfoDict[upc]
        for key in sorted(shippingInfos.keys()):
            for i in range(key-1, terminateIndex, -1):
                shippingInfos[i] = copy.deepcopy( shippingInfos[key] )
                shippingInfos[i].pop('preparation', None)
                shippingInfos[i].pop('ASIN', None)
            terminateIndex = key
        return shippingInfos
    return {}

def addOrModifyPackInfo(packsInfo, packInfoToInsert):
    packAmt = packInfoToInsert['packAmt']
    if packAmt == None:
        return
    packAmt = int(packAmt)

    if packAmt not in packsInfo:
        packsInfo[packAmt] = {}

    if packInfoToInsert != None and 'ASIN' in packInfoToInsert:
        packsInfo[packAmt]['ASIN'] = packInfoToInsert['ASIN']
    
    if packInfoToInsert['preparation'] != None:
        try:
            if 'preparation' in packsInfo[packAmt]:
                packsInfo[packAmt]['preparation'] = packsInfo[packAmt]['preparation'] + ' && ' + packInfoToInsert['preparation']
            else:
                packsInfo[packAmt]['preparation'] = packInfoToInsert['preparation']
        except (TypeError):
            pass

def updateProductInfo(curProductInfo, oldProductInfo):
    oldProductInfo['stockNo'] = curProductInfo['stockNo']
    oldProductInfo['shelfLocation'] = curProductInfo['shelfLocation']
    oldProductInfo['costPerBox'] = curProductInfo['costPerBox']
    oldProductInfo['quantityPerBox'] = curProductInfo['quantityPerBox']

def updateCheaperProductInfoIfNeeded(curProductInfo, oldProductInfo):
    try:
        curCostPerBox = curProductInfo['costPerBox']/curProductInfo['quantityPerBox']
        oldCostPerBox = oldProductInfo['costPerBox']/oldProductInfo['quantityPerBox']
        # Update existing info if cheaper. Carefully write over oldProductInfo to avoid overwritting past data
        if curCostPerBox != oldCostPerBox:
            print(  str.format("Price change detected:\nOld: {}\nNew: {}\n"
            , str(oldProductInfo), str(curProductInfo) )  )
            if curCostPerBox < oldCostPerBox:
                print("Old product info was replaced due to cheaper price.")
                updateProductInfo(curProductInfo, oldProductInfo)
    except TypeError:
        pass

def placeProductToInsert(userId:ObjectId, upcToShippingInfoDict:dict, dictOfProdsToInsert:dict, productToInsert:dict, packInfoToInsert:dict, sheetTitle:str):
    if 'packAmt' not in packInfoToInsert:
        return
    if 'UPC' in productToInsert:
        upc = productToInsert['UPC']
        #If the product info not in dict, add. Else, check/update if product cost is cheaper (Different wholesale might sell cheaper)
        if upc not in dictOfProdsToInsert:
            productToInsert['userId'] = userId
            productToInsert['wholesaleComp'] = sheetTitle
            productToInsert['packsInfo'] = getShippingInfo(upcToShippingInfoDict, upc)
            dictOfProdsToInsert[upc] = productToInsert
        else:
            updateCheaperProductInfoIfNeeded( productToInsert, dictOfProdsToInsert[upc] )

        addOrModifyPackInfo(dictOfProdsToInsert[upc]['packsInfo'], packInfoToInsert)
    elif len(productToInsert) > 0:
        print(str.format('UPC not found in: {}', productToInsert) )

def insertAllValidRowsToProductDict(userId:ObjectId, upcToShippingInfoDict:dict, sheet, headerRow
    , mainHeadersToMongoNameDict:dict, colNumToHeaderNameDict:dict
    , packInfoHeadersToMongoNameDict, packInfoColNumToHeaderNameDict, dictOfProdsToInsert)->None:
    for row in sheet.iter_rows(row_offset=headerRow):
        productToInsert = getDictMongoHeaderToCellVal(row, mainHeadersToMongoNameDict
                                             , colNumToHeaderNameDict)

        if productToInsert == None or 'UPC' not in productToInsert:
            continue
        upc = productToInsert['UPC']
        if upc != None and upc.lower() != 'missing' and upc.lower() != 'ignore':
            packInfoToInsert = getDictMongoHeaderToCellVal(row, packInfoHeadersToMongoNameDict
                            , packInfoColNumToHeaderNameDict)
        
            placeProductToInsert(userId, upcToShippingInfoDict, dictOfProdsToInsert
                            , productToInsert, packInfoToInsert, sheet.title)
        
def processSheet(userId:ObjectId, sheet, upcToShippingInfoDict:dict, headerRow:int
, mainHeadersToMongoNameDict:dict, packInfoHeadersToMongoNameDict:dict, dictOfProdsToInsert:dict):
    print( str.format('==================== Reading: {} ====================', sheet.title) )
    
    mainColNumToHeaderNameDict = getColNumToHeaderNameDict(sheet, headerRow, mainHeadersToMongoNameDict)
    packInfoColNumToHeaderNameDict = getColNumToHeaderNameDict(sheet, headerRow, packInfoHeadersToMongoNameDict)
    insertAllValidRowsToProductDict(userId, upcToShippingInfoDict, sheet, headerRow
    , mainHeadersToMongoNameDict, mainColNumToHeaderNameDict
    , packInfoHeadersToMongoNameDict, packInfoColNumToHeaderNameDict, dictOfProdsToInsert)

def getFirstExcelPathFromFolder(folderPath):
    for file in os.listdir(folderPath):
        if file.endswith('xlsx'):
            return os.path.join(folderPath, file)
    return None

def getNameToIdDict(db, userId):
    shipCollection = db['shippings']
    shipNameToIdDict = {}   

    for shipMethod in shipCollection.find({'userId': ObjectId(userId)},{ '_id': 1, 'shipCompanyName':1, 'shipMethodName':1}):
        key = str.format('{} - {}', shipMethod['shipCompanyName'], shipMethod['shipMethodName'])
        shipNameToIdDict[key] = shipMethod['_id']
    return shipNameToIdDict

def addToUpcToShippingInfoDict(upcToShippingInfoDict, shipNameToIdDict, row):
    upc = row[0].value
    if upc != None:
        packAmt = int(row[2].value)
        shipType = row[3].value
        ozWeight = row[4].value
        packaging = row[5].value
        preparation = row[6].value
        if upc not in upcToShippingInfoDict:
            upcToShippingInfoDict[upc] = {}
        upcToShippingInfoDict[upc][packAmt] = {'shipMethodId':shipNameToIdDict[shipType]
                                                , 'ozWeight':ozWeight, 'packaging':packaging, 'preparation':preparation}

def getUpcToShippingInfoDict(shipMethodExcelPath, db, userId):
    upcToShippingInfoDict = {}
    shipNameToIdDict = getNameToIdDict(db, userId)

    if shipMethodExcelPath == None:
        return upcToShippingInfoDict
    wb = openpyxl.load_workbook(shipMethodExcelPath, data_only=True)
    for sheet in wb:
        for row in sheet.iter_rows(row_offset=1):
            addToUpcToShippingInfoDict(upcToShippingInfoDict, shipNameToIdDict, row)

    return upcToShippingInfoDict

def getWholesaleExcelPath(rootFolderName, wholesaleFolderName):
    wholesaleFolderPath = os.path.join(rootFolderName, wholesaleFolderName)
    return getFirstExcelPathFromFolder(wholesaleFolderPath)

def getShipMethodExcelPath(rootFolderName, shipMethodFolderName):
    shipMethodFolderPath = os.path.join(rootFolderName, shipMethodFolderName)
    return getFirstExcelPathFromFolder(shipMethodFolderPath)

def getUserId(db):
    userCollection = db['users']
    while True:
        userEmail = input('Enter existing email: ')
        queryDoc = userCollection.find_one( {'email': userEmail} )
        if queryDoc != None:
            return (userEmail, ObjectId(queryDoc['_id']) )
        print('Invalid email... Try again...')

def getConfirmation(userEmail):
    while True:
        confirm = input( str.format("Proceeding will delete all products of {}. Proceed? (Y)es or (N)o: ", userEmail) )
        confirm = confirm.strip().upper()
        if confirm == 'Y' or confirm == 'N':
            return confirm
        print('Invalid response... Try again...')

def insertToMongoDb(prodCollection, dictOfProdsToInsert):
    print("Inserting to database...")
    if len(dictOfProdsToInsert) > 0:
        listToInsert = []
        for prodToInsert in dictOfProdsToInsert.values():
            packInfoAsList = []
            packsInfo = prodToInsert['packsInfo']
            for key in sorted(packsInfo.keys()):
                packInfo = packsInfo[key]
                packInfo['packAmt'] = key
                packInfoAsList.append(packInfo)
            prodToInsert['packsInfo'] = packInfoAsList
            listToInsert.append(prodToInsert)

        prodCollection.insert_many(listToInsert)

def promptUserEmailAndIntegrateFiles(db, wholesaleExcelPath, shipMethodExcelPath):
    mainHeadersToMongoNameDict = {'UPC':'UPC', 'product name':'name', 'stock no':'stockNo'
                                   , 'location':'shelfLocation', 'total cost':'costPerBox'
                                   , 'box amount':'quantityPerBox'}
    packInfoHeadersToMongoNameDict = {'pack':'packAmt', 'ASIN':'ASIN', 'prep':'preparation'}   
    email, userId = getUserId(db)
    confirmation = getConfirmation(email)
    if confirmation == 'N':
        return
    print( str.format("Deleting all products for {}...", email) )
    prodCollection = db['products']
    prodCollection.delete_many({'userId': userId})

    print("Reading excel files...")
    upcToShippingInfoDict = getUpcToShippingInfoDict(shipMethodExcelPath, db, userId)
        
    wb = openpyxl.load_workbook(wholesaleExcelPath, data_only=True)
    headerRow = 2

    dictOfProdsToInsert = {}
    for sheet in wb:
        processSheet(userId, sheet, upcToShippingInfoDict, headerRow, mainHeadersToMongoNameDict, packInfoHeadersToMongoNameDict, dictOfProdsToInsert)
    insertToMongoDb(prodCollection, dictOfProdsToInsert)

if __name__ == '__main__':
    rootFolderName = os.path.dirname(os.path.abspath(__file__))
    wholesaleFolderName = 'placeWholesaleExcelFileHere'
    wholesaleExcelPath = getWholesaleExcelPath(rootFolderName, wholesaleFolderName)
    shipMethodExcelPath = getShipMethodExcelPath(rootFolderName, 'placeShipMethodExcelFileHere')
    if wholesaleExcelPath == None:
        print( str.format('Error: Excel file not found in: {}', wholesaleFolderName) )
    else:
        try: 
            client = pymongo.MongoClient() #This connects to 'localhost', port# 27017 by default
            db = client['inventory-manager']
            print("Connected to mongodb successfully!")
            promptUserEmailAndIntegrateFiles(db, wholesaleExcelPath, shipMethodExcelPath)
            print("Successfully inserted to database...")
        except pymongo.errors.ConnectionFailure as e:
            print(e)
    input('\nPress enter to terminate program...')